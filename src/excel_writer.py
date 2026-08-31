from __future__ import annotations

import argparse
import json
import os
import re
import shutil
import unicodedata
import uuid
from copy import copy
from datetime import datetime
from pathlib import Path
from typing import Any


DEFAULT_FINAL_RECORDS = Path("data/cache/final_records.json")
DEFAULT_OVERRIDE_SUMMARY = Path("data/cache/override_summary.json")
DEFAULT_OVERRIDE_DIFF = Path("data/diagnostics/override_diff.json")

HEADERS = ["序号", "客户", "项目名称", "项目内容", "预算", "开标时间", "中标厂商", "中标金额", "备注", "来源链接"]
COLUMN_FIELDS = [
    "serial", "customer", "project_name", "content", "budget", "bid_open_time",
    "winner", "award_amount", "note", "source_url",
]
BLOCKING_OVERRIDE_SUMMARY_KEYS = (
    "invalid_override_count",
    "ambiguous_override_count",
    "conflict_override_count",
)
BLOCKING_DIFF_STATUSES = {"invalid", "ambiguous", "conflict"}
FORMULA_ERRORS = {"#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A"}


class WorkbookContractError(RuntimeError):
    """The workbook shape conflicts with the stable Excel field contract."""


def timestamp() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")


def normalize_text(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def normalize_key_part(value: object) -> str:
    normalized = unicodedata.normalize("NFKC", normalize_text(value)).lower()
    return re.sub(r"[\s（）()【】\[\]《》<>“”\"':：;；,，.。_\-—+]", "", normalized)


def record_key(record: dict[str, Any]) -> str:
    detail_id = normalize_key_part(record.get("award_detail_id"))
    if detail_id:
        return f"award_detail:{detail_id}"
    customer = normalize_key_part(record.get("customer"))
    project = normalize_key_part(record.get("project_name"))
    if customer or project:
        return f"{customer}|{project}"
    return normalize_key_part(record.get("source_file"))


def merge_record(base: dict[str, Any], incoming: dict[str, Any]) -> dict[str, Any]:
    merged = dict(base)
    for field in (
        "customer", "project_name", "content", "budget", "bid_open_time",
        "winner", "award_amount", "source_url",
    ):
        if not normalize_text(merged.get(field)) and normalize_text(incoming.get(field)):
            merged[field] = incoming[field]
    notes: list[str] = []
    for candidate in (merged.get("note"), incoming.get("note")):
        value = normalize_text(candidate)
        if value and value not in notes:
            notes.append(value)
    merged["note"] = "；".join(notes)
    return merged


def merge_input_records(records: list[dict[str, Any]]) -> list[dict[str, Any]]:
    merged: dict[str, dict[str, Any]] = {}
    for record in records:
        key = record_key(record)
        if not key:
            continue
        merged[key] = merge_record(merged[key], record) if key in merged else dict(record)
    return list(merged.values())


def note_for_record(record: dict[str, Any]) -> str:
    parts: list[str] = []
    if record.get("doc_type"):
        parts.append(f"类型：{record['doc_type']}")
    if record.get("note"):
        parts.append(normalize_text(record["note"]))
    if not record.get("text_chars"):
        parts.append("正文未能自动提取，已按文件名线索整理")
    return "；".join(part for part in parts if part)


def row_from_record(record: dict[str, Any], serial: object, *, write_source_url: bool) -> list[Any]:
    row: list[Any] = [
        serial,
        normalize_text(record.get("customer")),
        normalize_text(record.get("project_name")),
        normalize_text(record.get("content")),
        normalize_text(record.get("budget")),
        normalize_text(record.get("bid_open_time")),
        normalize_text(record.get("winner")),
        normalize_text(record.get("award_amount")),
        note_for_record(record),
    ]
    if write_source_url:
        row.append(normalize_text(record.get("source_url")))
    return row


def build_payload(args: argparse.Namespace, temp_output: Path, summary_path: Path) -> dict[str, Any]:
    """Retain the historical payload helper for Python callers; no Node runtime consumes it."""
    return {
        "excelPath": str(Path(args.excel).resolve()).replace("\\", "/"),
        "recordsPath": str(Path(args.records).resolve()).replace("\\", "/"),
        "tempOutputPath": str(temp_output.resolve()).replace("\\", "/"),
        "summaryPath": str(summary_path.resolve()).replace("\\", "/"),
        "sheetName": args.sheet_name,
        "allowExcelRowId": bool(getattr(args, "allow_excel_row_id", False)),
        "appendSourceNotes": bool(getattr(args, "append_source_notes", False)),
        "writeSourceUrl": bool(getattr(args, "write_source_url", False)),
        "overwriteFields": list(getattr(args, "overwrite_fields", []) or []),
    }


def is_final_records_path(records_path: Path) -> bool:
    try:
        return records_path.resolve() == DEFAULT_FINAL_RECORDS.resolve()
    except FileNotFoundError:
        return records_path.as_posix().replace("\\", "/").endswith(DEFAULT_FINAL_RECORDS.as_posix())


def load_json(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8"))


def validate_final_record_inputs(args: argparse.Namespace) -> None:
    records_path = Path(args.records)
    if not is_final_records_path(records_path):
        return
    if not records_path.exists():
        raise FileNotFoundError(f"FinalRecord file not found: {records_path}")

    summary_path = Path(args.override_summary)
    if not summary_path.exists():
        raise FileNotFoundError(f"Override summary file not found: {summary_path}")
    summary = load_json(summary_path)
    blocking_summary = {key: int(summary.get(key, 0) or 0) for key in BLOCKING_OVERRIDE_SUMMARY_KEYS}
    blocking_summary = {key: value for key, value in blocking_summary.items() if value > 0}
    if blocking_summary:
        raise RuntimeError(f"Blocking override summary counts before Excel write: {blocking_summary}")

    diff_path = Path(args.override_diff)
    if diff_path.exists():
        diff_items = load_json(diff_path)
        if not isinstance(diff_items, list):
            raise ValueError(f"Expected override diff JSON array: {diff_path}")
        blocking = [item for item in diff_items if item.get("match_status") in BLOCKING_DIFF_STATUSES]
        if blocking:
            raise RuntimeError(f"Blocking override diff statuses before Excel write: {len(blocking)}")


def _load_openpyxl() -> tuple[Any, type[Exception], type[Exception]]:
    try:
        from openpyxl import load_workbook
        from openpyxl.utils.exceptions import InvalidFileException
        from zipfile import BadZipFile
    except ImportError as exc:
        raise RuntimeError("处理 XLSX 需要安装 openpyxl，请先安装项目依赖。") from exc
    return load_workbook, InvalidFileException, BadZipFile


def _last_value_row(sheet: Any, last_column: int) -> int:
    for row_number in range(sheet.max_row, 0, -1):
        if any(normalize_text(sheet.cell(row_number, column).value) for column in range(1, last_column + 1)):
            return row_number
    return 0


def _validate_headers(sheet: Any, *, write_source_url: bool) -> None:
    current = [normalize_text(sheet.cell(1, column).value) for column in range(1, 11)]
    for index in range(8):
        if current[index] != HEADERS[index]:
            raise WorkbookContractError(
                f"Excel 数据契约冲突：第 {index + 1} 列表头 {current[index]!r}，预期 {HEADERS[index]!r}。"
            )
    if not current[8]:
        sheet.cell(1, 9).value = HEADERS[8]
    elif current[8] != HEADERS[8]:
        raise WorkbookContractError(f"Excel 数据契约冲突：第 9 列表头 {current[8]!r}，预期 {HEADERS[8]!r}。")
    if write_source_url and current[9] != HEADERS[9]:
        raise WorkbookContractError(f"Excel 数据契约冲突：第 10 列表头 {current[9]!r}，预期 {HEADERS[9]!r}。")


def _copy_row_format(sheet: Any, source_row: int, target_row: int, last_column: int) -> None:
    if source_row <= 0 or source_row == target_row:
        return
    for column in range(1, last_column + 1):
        source = sheet.cell(source_row, column)
        target = sheet.cell(target_row, column)
        if source.has_style:
            target._style = copy(source._style)
        if source.number_format:
            target.number_format = source.number_format
    source_dimension = sheet.row_dimensions[source_row]
    if source_dimension.height is not None:
        sheet.row_dimensions[target_row].height = source_dimension.height
    if source_dimension.hidden:
        sheet.row_dimensions[target_row].hidden = True


def _formula_error_ndjson(workbook: Any) -> str:
    errors: list[dict[str, str]] = []
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                value = cell.value
                if isinstance(value, str) and any(error in value for error in FORMULA_ERRORS):
                    errors.append({"sheet": sheet.title, "cell": cell.coordinate, "value": value})
                    if len(errors) >= 100:
                        return "\n".join(json.dumps(item, ensure_ascii=False) for item in errors)
    return "\n".join(json.dumps(item, ensure_ascii=False) for item in errors)


def _workbook_summary(workbook: Any, sheet: Any, records: list[dict[str, Any]], args: argparse.Namespace) -> dict[str, Any]:
    write_source_url = bool(getattr(args, "write_source_url", False))
    allow_excel_row_id = bool(getattr(args, "allow_excel_row_id", False))
    append_source_notes = bool(getattr(args, "append_source_notes", False))
    overwrite_fields = set(getattr(args, "overwrite_fields", []) or [])
    is_final_records = is_final_records_path(Path(args.records))
    last_column = 10 if write_source_url else 9
    _validate_headers(sheet, write_source_url=write_source_url)
    existing_last_row = _last_value_row(sheet, 10)
    existing_row_count = max(existing_last_row - 1, 0)

    existing_index: dict[str, int] = {}
    blank_customer_index: dict[str, list[int]] = {}
    max_serial = 0
    for row_number in range(2, existing_last_row + 1):
        values = [sheet.cell(row_number, column).value for column in range(1, 11)]
        try:
            max_serial = max(max_serial, int(float(values[0])))
        except (TypeError, ValueError):
            pass
        key = f"{normalize_key_part(values[1])}|{normalize_key_part(values[2])}"
        if (normalize_key_part(values[1]) or normalize_key_part(values[2])) and key not in existing_index:
            existing_index[key] = row_number
        project_key = normalize_key_part(values[2])
        if project_key and not normalize_key_part(values[1]):
            blank_customer_index.setdefault(project_key, []).append(row_number)

    matched_by = {
        "excel_row_id": 0, "source_file": 0, "exact_key": 0,
        "project_name_blank_customer": 0, "appended": 0, "ambiguous": 0,
    }
    updated_rows = 0
    updated_cells = 0
    overwritten_cells = 0
    overwrite_details: list[dict[str, Any]] = []
    ambiguous_details: list[dict[str, Any]] = []
    outcomes: list[dict[str, Any]] = []
    appended_count = 0
    merged_records = merge_input_records(records)

    def source_file_match(record: dict[str, Any]) -> dict[str, Any] | None:
        source_key = normalize_key_part(record.get("source_file"))
        if not source_key:
            return None
        rows = [
            row_number for row_number in range(2, existing_last_row + 1)
            if source_key in normalize_key_part(sheet.cell(row_number, 9).value)
        ]
        if len(rows) == 1:
            return {"row": rows[0], "matched_by": "source_file"}
        if len(rows) > 1:
            return {"ambiguous": True, "matched_by": "source_file", "rows": rows}
        return None

    def find_match(record: dict[str, Any]) -> dict[str, Any] | None:
        if is_final_records or allow_excel_row_id:
            try:
                excel_row = int(record.get("excel_row_id") or 0)
            except (TypeError, ValueError):
                excel_row = 0
            if 2 <= excel_row <= existing_last_row:
                return {"row": excel_row, "matched_by": "excel_row_id"}
        key = record_key(record)
        if not normalize_key_part(record.get("award_detail_id")) and key and key in existing_index:
            return {"row": existing_index[key], "matched_by": "exact_key"}
        if is_final_records:
            match = source_file_match(record)
            if match:
                return match
        if is_final_records and normalize_key_part(record.get("customer")):
            rows = blank_customer_index.get(normalize_key_part(record.get("project_name")), [])
            if len(rows) == 1:
                return {"row": rows[0], "matched_by": "project_name_blank_customer"}
            if len(rows) > 1:
                return {"ambiguous": True, "matched_by": "project_name_blank_customer", "rows": rows}
        return None

    for record in merged_records:
        match = find_match(record)
        if match and match.get("ambiguous"):
            detail = {
                "matched_by": match["matched_by"],
                "source_document_id": record.get("source_document_id", ""),
                "source_file": record.get("source_file", ""),
                "project_name": record.get("project_name", ""),
                "candidate_rows": match["rows"],
            }
            ambiguous_details.append(detail)
            matched_by["ambiguous"] += 1
            outcomes.append({
                "source_document_id": record.get("source_document_id", ""),
                "award_detail_id": record.get("award_detail_id", ""),
                "status": "failed", "action": "conflict", "excel_row": 0,
                "reason": "ambiguous_match",
            })
            continue

        if match and match.get("row"):
            excel_row = int(match["row"])
            matched_by[str(match["matched_by"])] += 1
            current = [sheet.cell(excel_row, column).value for column in range(1, last_column + 1)]
            candidate = row_from_record(record, current[0], write_source_url=write_source_url)
            updates: list[int] = []
            for index in range(1, len(candidate)):
                current_text = normalize_text(current[index])
                candidate_text = normalize_text(candidate[index])
                if not current_text and candidate_text:
                    current[index] = candidate[index]
                    updates.append(index)
                elif COLUMN_FIELDS[index] in overwrite_fields and candidate_text and current_text != candidate_text:
                    overwrite_details.append({
                        "excel_row": excel_row, "field": COLUMN_FIELDS[index],
                        "old_value": current_text, "new_value": candidate_text,
                    })
                    current[index] = candidate[index]
                    updates.append(index)
                    overwritten_cells += 1
                elif index == 8 and append_source_notes and current_text and candidate_text and candidate_text not in current_text:
                    current[index] = f"{current_text}；{candidate_text}"
                    updates.append(index)
            for index in updates:
                sheet.cell(excel_row, index + 1).value = current[index]
            if updates:
                updated_rows += 1
                updated_cells += len(updates)
            outcomes.append({
                "source_document_id": record.get("source_document_id", ""),
                "award_detail_id": record.get("award_detail_id", ""),
                "status": "written", "action": "update" if updates else "unchanged",
                "excel_row": excel_row, "reason": "",
            })
            continue

        max_serial += 1
        excel_row = existing_last_row + appended_count + 1
        template_row = existing_last_row if existing_last_row >= 2 else (2 if sheet.max_row >= 2 else 1)
        _copy_row_format(sheet, template_row, excel_row, last_column)
        values = row_from_record(record, max_serial, write_source_url=write_source_url)
        for index, value in enumerate(values, 1):
            sheet.cell(excel_row, index).value = value
        appended_count += 1
        matched_by["appended"] += 1
        outcomes.append({
            "source_document_id": record.get("source_document_id", ""),
            "award_detail_id": record.get("award_detail_id", ""),
            "status": "written", "action": "insert", "excel_row": excel_row, "reason": "",
        })

    final_last_row = _last_value_row(sheet, 10)
    seen: set[str] = set()
    duplicate_keys = 0
    for row_number in range(2, final_last_row + 1):
        key = f"{normalize_key_part(sheet.cell(row_number, 2).value)}|{normalize_key_part(sheet.cell(row_number, 3).value)}"
        if key == "|":
            continue
        if key in seen:
            duplicate_keys += 1
        seen.add(key)

    return {
        "inputRecords": len(records),
        "mergedRecords": len(merged_records),
        "existingRows": existing_row_count,
        "appendedRows": appended_count,
        "updatedRows": updated_rows,
        "updatedCells": updated_cells,
        "overwrittenCells": overwritten_cells,
        "overwriteDetails": overwrite_details,
        "matched_by": matched_by,
        "updated_existing_rows": updated_rows,
        "appended_rows": appended_count,
        "ambiguous_matches": len(ambiguous_details),
        "skipped_due_to_ambiguous_match": len(ambiguous_details),
        "ambiguous_match_details": ambiguous_details,
        "record_outcomes": outcomes,
        "finalRows": max(final_last_row - 1, 0),
        "duplicateKeys": duplicate_keys,
        "formulaErrors": _formula_error_ndjson(workbook),
    }


def _load_workbook_for_write(excel_path: Path) -> Any:
    load_workbook, invalid_file, bad_zip = _load_openpyxl()
    try:
        return load_workbook(excel_path, read_only=False, data_only=False)
    except PermissionError as exc:
        raise RuntimeError(f"Excel 文件被占用或无读取权限：{excel_path}") from exc
    except (invalid_file, bad_zip, EOFError) as exc:
        raise RuntimeError(f"Excel 文件损坏或不是有效的 XLSX：{excel_path}") from exc
    except OSError as exc:
        raise RuntimeError(f"无法读取 Excel 文件：{excel_path}（{exc}）") from exc


def run_writer(args: argparse.Namespace) -> dict[str, Any]:
    excel_path = Path(args.excel)
    records_path = Path(args.records)
    if not excel_path.exists():
        raise FileNotFoundError(f"Excel file not found: {excel_path}")
    if not records_path.exists():
        raise FileNotFoundError(f"Record file not found: {records_path}")
    validate_final_record_inputs(args)

    raw_records = load_json(records_path)
    if not isinstance(raw_records, list) or not all(isinstance(item, dict) for item in raw_records):
        raise ValueError(f"Expected record JSON array: {records_path}")
    records = [dict(item) for item in raw_records]
    summary_path = Path(args.summary)
    summary_path.parent.mkdir(parents=True, exist_ok=True)
    backup_dir = Path(args.backup_dir)
    backup_dir.mkdir(parents=True, exist_ok=True)

    workbook = _load_workbook_for_write(excel_path)
    try:
        if args.sheet_name not in workbook.sheetnames:
            raise WorkbookContractError(f"Excel 数据契约冲突：工作表不存在：{args.sheet_name}")
        summary = _workbook_summary(workbook, workbook[args.sheet_name], records, args)

        backup_path = backup_dir / f"{excel_path.stem}_backup_{timestamp()}{excel_path.suffix}"
        try:
            shutil.copy2(excel_path, backup_path)
        except PermissionError as exc:
            raise RuntimeError(f"无法创建 Excel 备份，目录无写入权限：{backup_dir}") from exc
        except OSError as exc:
            raise RuntimeError(f"无法创建 Excel 备份：{backup_path}（{exc}）") from exc

        temp_output = excel_path.with_name(f".{excel_path.stem}.tmp.{uuid.uuid4().hex}{excel_path.suffix}")
        try:
            workbook.save(temp_output)
            if not temp_output.exists() or temp_output.stat().st_size == 0:
                raise RuntimeError("Excel 临时文件保存失败：未生成有效输出。")
            try:
                os.replace(temp_output, excel_path)
            except PermissionError as exc:
                raise RuntimeError(f"Excel 文件被占用或无替换权限，原文件未修改：{excel_path}") from exc
            except OSError as exc:
                raise RuntimeError(f"Excel 原子替换失败，原文件未修改：{excel_path}（{exc}）") from exc
        except PermissionError as exc:
            raise RuntimeError(f"Excel 临时文件保存失败，目录无写入权限：{excel_path.parent}") from exc
        except OSError as exc:
            raise RuntimeError(f"Excel 临时文件保存失败，原文件未修改：{excel_path}（{exc}）") from exc
        finally:
            if temp_output.exists():
                try:
                    temp_output.unlink()
                except OSError:
                    pass
    finally:
        workbook.close()

    summary["backupPath"] = str(backup_path)
    summary["recordsPath"] = str(records_path)
    summary["stdout"] = json.dumps(summary, ensure_ascii=False)
    summary_path.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    return summary


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Write ExtractedRecord data into 招投标.xlsx safely and idempotently.")
    parser.add_argument("--records", default="data/cache/extracted_records.json")
    parser.add_argument("--excel", default="招投标.xlsx")
    parser.add_argument("--sheet-name", default="Sheet1")
    parser.add_argument("--backup-dir", default="data/backups")
    parser.add_argument("--summary", default="data/cache/excel_writer_summary.json")
    parser.add_argument("--override-summary", default=str(DEFAULT_OVERRIDE_SUMMARY))
    parser.add_argument("--override-diff", default=str(DEFAULT_OVERRIDE_DIFF))
    return parser.parse_args()


def main() -> None:
    summary = run_writer(parse_args())
    print(
        "excel_writer completed "
        f"appended={summary['appendedRows']} updated={summary['updatedRows']} backup={summary['backupPath']}"
    )


if __name__ == "__main__":
    main()
