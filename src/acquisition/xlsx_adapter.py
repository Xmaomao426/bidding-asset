"""Conservative XLSX-to-existing-candidate adapter for Acquisition Inbox."""

from __future__ import annotations

import hashlib
from dataclasses import dataclass
from datetime import date, datetime, time, timezone
from functools import lru_cache
from pathlib import Path
from typing import Any

from src.parser import file_hash
from src.project_number import validated_project_number
from src.award_detail import award_detail_id, business_group_id

from .xlsx_fields import BUSINESS_FIELDS, IDENTITY_FIELDS, mapped_field, mapping_source


DEFAULT_MAX_XLSX_VALID_ROWS = 2000
MAX_HEADER_SCAN_ROWS = 20
MAX_HEADER_SCAN_COLUMNS = 200
MAX_REPORTED_ROW_DIAGNOSTICS = 50
FORMAL_WORKBOOK_NAME = "招投标.xlsx"
SUMMARY_LABELS = {"合计", "总计", "汇总", "说明", "备注"}


class XlsxIngestionError(ValueError):
    """User-facing workbook failure with a safe structured diagnostic payload."""

    def __init__(self, message: str, processing_result: dict[str, Any] | None = None) -> None:
        super().__init__(message)
        self.processing_result = dict(processing_result or {})


@lru_cache(maxsize=1)
def openpyxl_api() -> tuple[Any, Any]:
    """Load the optional XLSX dependency only when an XLSX is processed."""
    try:
        from openpyxl import load_workbook as openpyxl_load_workbook
        from openpyxl.utils import get_column_letter as openpyxl_get_column_letter
    except ModuleNotFoundError as exc:
        if str(exc.name or "").split(".", 1)[0] != "openpyxl":
            raise
        raise XlsxIngestionError(
            "当前环境缺少 XLSX 解析依赖 openpyxl；其他文件和 URL 功能仍可使用。"
            "请使用项目 .venv 安装 requirements.txt 后重试该 XLSX。"
        ) from exc
    return openpyxl_load_workbook, openpyxl_get_column_letter


def load_workbook(*args: Any, **kwargs: Any) -> Any:
    loader, _column_letter = openpyxl_api()
    return loader(*args, **kwargs)


def get_column_letter(column: int) -> str:
    _loader, column_letter = openpyxl_api()
    return str(column_letter(column))


@dataclass(frozen=True)
class HeaderInfo:
    row_number: int
    column_count: int
    headers: tuple[str, ...]
    field_columns: dict[str, int]
    recognized: tuple[dict[str, Any], ...]


def parse_xlsx_workbook(
    path: Path,
    *,
    max_valid_rows: int = DEFAULT_MAX_XLSX_VALID_ROWS,
) -> dict[str, Any]:
    """Read one XLSX without saving it and return row payloads plus diagnostics."""
    source_path = Path(path)
    reject_formal_workbook(source_path)
    digest = file_hash(source_path)
    base = processing_base(source_path, digest)
    try:
        workbook = load_workbook(source_path, read_only=True, data_only=True, keep_links=False)
    except XlsxIngestionError:
        raise
    except Exception as exc:
        raise XlsxIngestionError(
            "无法读取 XLSX 工作簿；文件可能损坏、加密或格式不正确。",
            {**base, "parse_status": "failed", "extract_status": "failed", "parse_error": "工作簿无法读取。"},
        ) from exc

    sheet_results: list[dict[str, Any]] = []
    row_payloads: list[dict[str, Any]] = []
    failed_row_payloads: list[dict[str, Any]] = []
    skipped_rows: list[dict[str, Any]] = []
    failed_rows: list[dict[str, Any]] = []
    scanned_data_rows = 0
    skipped_row_count = 0
    valid_row_count = 0
    recognized_sheet_count = 0
    sheet_names = list(workbook.sheetnames)

    try:
        for worksheet in workbook.worksheets:
            header, sheet_nonempty, header_error = find_header(worksheet)
            if header is None:
                if not sheet_nonempty:
                    sheet_results.append(sheet_result(worksheet.title, "skipped_empty", reason="工作表为空。"))
                else:
                    sheet_results.append(sheet_result(
                        worksheet.title,
                        "failed",
                        reason=header_error or "未识别到普通单行表头。",
                    ))
                continue

            recognized_sheet_count += 1
            sheet_scanned = 0
            sheet_skipped = 0
            sheet_failed = 0
            sheet_valid = 0
            for row_number, values in enumerate(
                worksheet.iter_rows(
                    min_row=header.row_number + 1,
                    max_col=header.column_count,
                    values_only=True,
                ),
                start=header.row_number + 1,
            ):
                cleaned = [clean_cell_value(value) for value in values]
                if not any(cleaned):
                    skipped_row_count += 1
                    sheet_skipped += 1
                    add_row_diagnostic(skipped_rows, worksheet.title, row_number, "全空行。")
                    continue
                scanned_data_rows += 1
                sheet_scanned += 1
                try:
                    payload, reason = build_row_payload(
                        source_path,
                        digest,
                        worksheet.title,
                        row_number,
                        header,
                        cleaned,
                    )
                    if payload is None:
                        skipped_row_count += 1
                        sheet_skipped += 1
                        add_row_diagnostic(skipped_rows, worksheet.title, row_number, reason)
                        continue
                    valid_row_count += 1
                    sheet_valid += 1
                    if len(row_payloads) <= max_valid_rows:
                        row_payloads.append(payload)
                except Exception:
                    sheet_failed += 1
                    add_row_diagnostic(failed_rows, worksheet.title, row_number, "该行无法转换为业务候选。")
                    failed_row_payloads.append(failed_row_payload(
                        source_path,
                        digest,
                        worksheet.title,
                        row_number,
                        header,
                        cleaned,
                    ))

            sheet_results.append(sheet_result(
                worksheet.title,
                "parsed",
                header_row=header.row_number,
                headers=list(header.headers),
                recognized_headers=list(header.recognized),
                scanned_data_rows=sheet_scanned,
                successful_candidate_count=sheet_valid,
                skipped_row_count=sheet_skipped,
                failed_row_count=sheet_failed,
            ))
    finally:
        workbook.close()

    summary = {
        **base,
        "sheet_count": len(sheet_names),
        "sheet_results": sheet_results,
        "scanned_data_rows": scanned_data_rows,
        "successful_candidate_count": valid_row_count,
        "skipped_row_count": skipped_row_count,
        "failed_row_count": sum(int(result.get("failed_row_count") or 0) for result in sheet_results),
        "skipped_rows": skipped_rows,
        "failed_rows": failed_rows,
        "max_valid_rows": max_valid_rows,
    }
    if recognized_sheet_count == 0:
        raise XlsxIngestionError(
            "工作簿没有可识别的普通单行表头。",
            {**summary, "parse_status": "failed", "extract_status": "failed", "parse_error": "无可识别表头。"},
        )
    if valid_row_count > max_valid_rows:
        raise XlsxIngestionError(
            f"工作簿包含 {valid_row_count} 个有效数据行，超过同步处理上限 {max_valid_rows}；未创建任何行任务。",
            {**summary, "parse_status": "failed", "extract_status": "failed", "parse_error": "有效数据行超过同步上限。"},
        )
    if not row_payloads:
        raise XlsxIngestionError(
            "工作簿没有满足项目名称、有效项目编号或来源链接条件的有效数据行。",
            {**summary, "parse_status": "failed", "extract_status": "failed", "parse_error": "无有效数据行。"},
        )

    try:
        enrich_hyperlinks_and_validate_headers(source_path, row_payloads, sheet_results)
    except XlsxIngestionError as exc:
        if not exc.processing_result:
            exc.processing_result = {
                **summary,
                "parse_status": "failed",
                "extract_status": "failed",
                "parse_error": str(exc),
            }
        raise
    annotate_business_groups(row_payloads, summary)
    summary.update({"parse_status": "success", "extract_status": "success", "parse_error": ""})
    return {"summary": summary, "rows": row_payloads, "failed_rows": failed_row_payloads}


def reject_formal_workbook(path: Path) -> None:
    if Path(path).name.casefold() == FORMAL_WORKBOOK_NAME.casefold():
        raise XlsxIngestionError(
            "根目录正式《招投标.xlsx》不允许作为普通资料采集来源；该文件属于未来 Historical Asset Import 范围。",
            {
                **processing_base(Path(path), ""),
                "parse_status": "failed",
                "extract_status": "failed",
                "parse_error": "正式业务工作簿禁止普通采集。",
            },
        )


def find_header(worksheet: Any) -> tuple[HeaderInfo | None, bool, str]:
    max_row = int(worksheet.max_row or 0)
    max_column = min(int(worksheet.max_column or 0), MAX_HEADER_SCAN_COLUMNS)
    sheet_nonempty = False
    candidates: list[HeaderInfo] = []
    for row_number, row in enumerate(
        worksheet.iter_rows(min_row=1, max_row=min(max_row, MAX_HEADER_SCAN_ROWS), max_col=max_column, values_only=True),
        start=1,
    ):
        values = [clean_cell_value(value) for value in row]
        if any(values):
            sheet_nonempty = True
        last_nonempty = max((index for index, value in enumerate(values, 1) if value), default=0)
        if not last_nonempty:
            continue
        headers = values[:last_nonempty]
        field_columns: dict[str, int] = {}
        recognized: list[dict[str, Any]] = []
        duplicate_field = ""
        for column, header in enumerate(headers, 1):
            field = mapped_field(header)
            if not field:
                continue
            if field in field_columns:
                duplicate_field = field
                break
            field_columns[field] = column
            recognized.append({
                "column": get_column_letter(column),
                "header": header,
                "field": field,
                "mapping_source": mapping_source(header),
            })
        if duplicate_field:
            return None, True, f"表头将多个列映射到同一字段 {duplicate_field}，需要人工整理。"
        is_single_identity_column = last_nonempty == 1 and bool(IDENTITY_FIELDS.intersection(field_columns))
        if IDENTITY_FIELDS.intersection(field_columns) and (len(field_columns) >= 2 or is_single_identity_column):
            candidates.append(HeaderInfo(row_number, last_nonempty, tuple(headers), field_columns, tuple(recognized)))
    if len(candidates) > 1:
        return None, sheet_nonempty, "检测到多个可能表头行；不支持复杂多行表头。"
    if candidates:
        return candidates[0], sheet_nonempty, ""
    return None, sheet_nonempty, "未识别到包含项目名称、项目编号或来源链接的精确表头。"


def build_row_payload(
    path: Path,
    digest: str,
    sheet_name: str,
    row_number: int,
    header: HeaderInfo,
    values: list[str],
) -> tuple[dict[str, Any] | None, str]:
    fields = {field: values[column - 1] if column <= len(values) else "" for field, column in header.field_columns.items()}
    fields = {key: str(value or "").strip() for key, value in fields.items()}
    fields["project_number"] = validated_project_number(fields.get("project_number"))
    fields.setdefault("project_name", "")
    fields.setdefault("source_url", "")
    if is_summary_row(fields):
        return None, "说明、汇总或备注行。"
    if not any(fields.get(field) for field in IDENTITY_FIELDS):
        return None, "缺少项目名称、有效项目编号和来源链接。"

    source_key = xlsx_row_source_key(digest, sheet_name, row_number)
    candidate_id = xlsx_candidate_id(source_key)
    raw_columns = [
        {
            "column": get_column_letter(column),
            "header": header.headers[column - 1],
            "value": values[column - 1] if column <= len(values) else "",
        }
        for column in range(1, header.column_count + 1)
        if header.headers[column - 1]
    ]
    business = {field: fields.get(field, "") for field in BUSINESS_FIELDS}
    is_tianyancha_xlsx = bool(set(header.headers).intersection({"发布日期", "省份地区", "招采单位", "招投标详情"}))
    source_trace = {
        "source_type": "xlsx_file_upload",
        "source_name": "xlsx_source_ingestion",
        "source_file": path.name,
        "original_file_name": path.name,
        "source_path": str(path),
        "file_sha256": digest,
        "sheet_name": sheet_name,
        "excel_row_number": row_number,
        "original_columns": raw_columns,
        "original_hyperlinks": [],
        "source_key": source_key,
        "is_tianyancha_xlsx": is_tianyancha_xlsx,
    }
    identity_payload = {**business, "file_type": ".xlsx", "file_hash": digest, "source_trace": source_trace}
    business["business_sequence"] = business.get("sequence", "")
    business["is_tianyancha_xlsx"] = is_tianyancha_xlsx
    business["business_group_id"] = business_group_id(identity_payload)
    business["award_detail_id"] = award_detail_id({**identity_payload, **business})
    source_trace.update({
        "business_sequence": business["business_sequence"],
        "publish_date": business.get("publish_date", ""),
        "business_group_id": business["business_group_id"],
        "award_detail_id": business["award_detail_id"],
    })
    candidate_row = {
        "candidate_id": candidate_id,
        "source_title": business.get("project_name") or business.get("source_url") or f"{path.name} / {sheet_name} / row {row_number}",
        "source_url": business.get("source_url", ""),
        "source_file": str(path),
        "discovered_time": utc_now(),
        "confidence": 0.0,
        **business,
        "source_trace": source_trace,
    }
    processing = {
        **business,
        "file_name": path.name,
        "file_type": ".xlsx",
        "source_path": str(path),
        "file_hash": digest,
        "sheet_name": sheet_name,
        "excel_row_number": row_number,
        "original_headers": list(header.headers),
        "original_columns": raw_columns,
        "original_hyperlinks": [],
        "parse_status": "success",
        "parse_error": "",
        "extract_status": "success",
    }
    return {
        "source_key": source_key,
        "candidate_row": candidate_row,
        "processing": processing,
        "sources": [{
            "source_type": "xlsx_file_upload",
            "source_file": path.name,
            "source_path": str(path),
            "file_sha256": digest,
            "sheet_name": sheet_name,
            "excel_row_number": row_number,
            "created_time": utc_now(),
            "status": "parsed",
        }],
    }, ""


def annotate_business_groups(row_payloads: list[dict[str, Any]], summary: dict[str, Any]) -> None:
    """Annotate rows and workbook summary without collapsing physical award-detail rows."""
    grouped: dict[str, list[dict[str, Any]]] = {}
    ungrouped = 0
    for payload in row_payloads:
        processing = dict(payload.get("processing") or {})
        group_id = str(processing.get("business_group_id") or "")
        if group_id:
            grouped.setdefault(group_id, []).append(payload)
        else:
            ungrouped += 1
    sizes = [len(rows) for rows in grouped.values()]
    summary.update({
        "project_count": len(grouped) + ungrouped,
        "award_detail_count": len(row_payloads),
        "multi_detail_project_count": sum(1 for size in sizes if size > 1),
        "max_award_details_per_project": max(sizes + ([1] if ungrouped else [0])),
    })
    for rows in grouped.values():
        detail_count = len(rows)
        for payload in rows:
            processing = payload["processing"]
            sequence = str(processing.get("business_sequence") or "")
            markers = {
                "project_group_label": f"业务序号 {sequence}（同项目共 {detail_count} 条中标明细）",
                "project_group_detail_count": detail_count,
                "is_multi_detail_project": detail_count > 1,
            }
            processing.update(markers)
            payload["candidate_row"].update(markers)
            payload["candidate_row"]["source_trace"].update(markers)


def enrich_hyperlinks_and_validate_headers(
    path: Path,
    row_payloads: list[dict[str, Any]],
    sheet_results: list[dict[str, Any]],
) -> None:
    try:
        workbook = load_workbook(path, read_only=False, data_only=True, keep_links=False)
    except XlsxIngestionError:
        raise
    except Exception as exc:
        raise XlsxIngestionError("工作簿值已读取，但无法安全提取单元格超链接。") from exc
    try:
        header_by_sheet = {
            str(result.get("sheet_name") or ""): result
            for result in sheet_results
            if result.get("status") == "parsed"
        }
        for sheet_name, result in header_by_sheet.items():
            worksheet = workbook[sheet_name]
            header_row = int(result.get("header_row") or 0)
            if any(cell_range.min_row <= header_row <= cell_range.max_row for cell_range in worksheet.merged_cells.ranges):
                raise XlsxIngestionError(f"工作表“{sheet_name}”使用合并表头；本轮只支持普通单行表头。")

        for payload in row_payloads:
            processing = payload["processing"]
            sheet_name = str(processing["sheet_name"])
            row_number = int(processing["excel_row_number"])
            worksheet = workbook[sheet_name]
            hyperlinks: list[dict[str, str]] = []
            source_url_target = ""
            any_target = ""
            for raw in processing.get("original_columns", []):
                column = str(raw.get("column") or "")
                if not column:
                    continue
                cell = worksheet[f"{column}{row_number}"]
                target = str(cell.hyperlink.target or "").strip() if cell.hyperlink and cell.hyperlink.target else ""
                if not target:
                    continue
                header = str(raw.get("header") or "")
                hyperlinks.append({"column": column, "header": header, "target": target})
                any_target = any_target or target
                if mapped_field(header) == "source_url":
                    source_url_target = source_url_target or target
            effective_url = source_url_target or any_target or str(processing.get("source_url") or "")
            processing["source_url"] = effective_url
            processing["original_hyperlinks"] = hyperlinks
            candidate = payload["candidate_row"]
            candidate["source_url"] = effective_url
            candidate["source_trace"]["source_url"] = effective_url
            candidate["source_trace"]["original_hyperlinks"] = hyperlinks
    finally:
        workbook.close()


def processing_base(path: Path, digest: str) -> dict[str, Any]:
    return {
        "file_name": Path(path).name,
        "file_type": Path(path).suffix.lower(),
        "source_path": str(path),
        "file_hash": digest,
        "doc_type": "XLSX 工作簿",
    }


def failed_row_payload(
    path: Path,
    digest: str,
    sheet_name: str,
    row_number: int,
    header: HeaderInfo,
    values: list[str],
) -> dict[str, Any]:
    source_key = xlsx_row_source_key(digest, sheet_name, row_number)
    raw_columns = [
        {
            "column": get_column_letter(column),
            "header": header.headers[column - 1],
            "value": values[column - 1] if column <= len(values) else "",
        }
        for column in range(1, header.column_count + 1)
        if header.headers[column - 1]
    ]
    processing = {
        **processing_base(path, digest),
        "sheet_name": sheet_name,
        "excel_row_number": row_number,
        "original_headers": list(header.headers),
        "original_columns": raw_columns,
        "original_hyperlinks": [],
        "parse_status": "success",
        "parse_error": "",
        "extract_status": "failed",
    }
    return {
        "source_key": source_key,
        "processing": processing,
        "sources": [{
            "source_type": "xlsx_file_upload",
            "source_file": path.name,
            "source_path": str(path),
            "file_sha256": digest,
            "sheet_name": sheet_name,
            "excel_row_number": row_number,
            "created_time": utc_now(),
            "status": "row_failed",
        }],
        "error_type": "XlsxRowError",
        "error_message": "该 Excel 行无法转换为业务候选。",
    }


def sheet_result(sheet_name: str, status: str, **values: Any) -> dict[str, Any]:
    return {"sheet_name": sheet_name, "status": status, **values}


def add_row_diagnostic(rows: list[dict[str, Any]], sheet_name: str, row_number: int, reason: str) -> None:
    if len(rows) < MAX_REPORTED_ROW_DIAGNOSTICS:
        rows.append({"sheet_name": sheet_name, "excel_row_number": row_number, "reason": reason})


def is_summary_row(fields: dict[str, str]) -> bool:
    project_name = str(fields.get("project_name") or "").strip()
    other_identity = bool(fields.get("project_number") or fields.get("source_url"))
    other_business = any(
        fields.get(field) for field in BUSINESS_FIELDS
        if field not in {"project_name", "project_number", "source_url", "note"}
    )
    normalized = project_name.rstrip("：:").strip()
    return bool(project_name and not other_identity and not other_business and (
        normalized in SUMMARY_LABELS or any(project_name.startswith(f"{label}：") for label in SUMMARY_LABELS)
    ))


def clean_cell_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat(sep=" ", timespec="seconds")
    if isinstance(value, date | time):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def xlsx_row_source_key(file_sha256: str, sheet_name: str, row_number: int) -> str:
    raw = f"{file_sha256}\n{sheet_name}\n{row_number}"
    return f"xlsx_row_sha256:{hashlib.sha256(raw.encode('utf-8')).hexdigest()}"


def xlsx_candidate_id(source_key: str) -> str:
    digest = hashlib.sha256(f"xlsx_candidate\n{source_key}".encode("utf-8")).hexdigest()[:16]
    return f"asset_candidate_{digest}"


def utc_now() -> str:
    return datetime.now(timezone.utc).isoformat(timespec="seconds").replace("+00:00", "Z")
