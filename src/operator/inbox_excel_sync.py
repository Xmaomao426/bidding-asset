"""Explicit Inbox-to-Excel adapter over the existing V1 Excel Writer."""

from __future__ import annotations

import json
import re
import unicodedata
import uuid
from argparse import Namespace
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from src.excel_writer import run_writer
from src.award_detail import award_detail_id, business_group_id, business_sequence
from src.matcher.record_matcher import project_numbers
from src.project_number import validated_project_number


NOT_WRITTEN = "NOT_WRITTEN"
WRITTEN = "WRITTEN"
CONFLICT = "CONFLICT"
FAILED = "FAILED"
NOT_APPLICABLE = "NOT_APPLICABLE"

INSERT = "insert"
UPDATE = "update"
UNCHANGED = "unchanged"
CONFLICT_ACTION = "conflict"

BUSINESS_FIELDS = (
    "customer",
    "project_name",
    "content",
    "budget",
    "bid_open_time",
    "winner",
    "award_amount",
    "source_url",
)
PROJECT_LEVEL_FIELDS = tuple(field for field in BUSINESS_FIELDS if field not in {"winner", "award_amount", "source_url"})


@dataclass(frozen=True)
class InboxExcelSyncPaths:
    excel: Path
    records_dir: Path
    summary_dir: Path
    backup_dir: Path
    runtime_dir: Path
    sheet_name: str = "Sheet1"


def sync_confirmed_project(
    *,
    inbox_id: str,
    extracted: dict[str, Any],
    project: dict[str, Any],
    source_file: str,
    source_url: str,
    confirmed_time: str,
    paths: InboxExcelSyncPaths,
) -> dict[str, Any]:
    """Preflight one confirmed project and invoke the existing writer only when safe."""
    record = build_excel_record(extracted, project, source_file, source_url, confirmed_time)
    if not normalize_text(record.get("project_name")):
        return sync_result(NOT_APPLICABLE, "", "项目名称缺失，未写入 Excel。")

    workbook = inspect_workbook(paths)
    audited_by_project, audited_by_detail = prior_excel_row_indexes(paths)
    plan = plan_record_excel_sync(record, workbook["rows"], audited_by_project, audited_by_detail)
    if plan["excel_status"] == CONFLICT:
        return plan

    matched_row = dict(plan.get("matched_row") or {})
    if matched_row:
        record["excel_row_id"] = int(matched_row["excel_row"])

    records_path = paths.records_dir / f"{safe_token(inbox_id)}.json"
    summary_path = paths.summary_dir / f"{safe_token(inbox_id)}.json"
    records_path.parent.mkdir(parents=True, exist_ok=True)
    summary_path.parent.mkdir(parents=True, exist_ok=True)
    records_path.write_text(json.dumps([record], ensure_ascii=False, indent=2), encoding="utf-8")

    action = str(plan["excel_action"])
    if action == UNCHANGED:
        summary = {
            "inputRecords": 1,
            "mergedRecords": 1,
            "existingRows": len(workbook["rows"]),
            "appendedRows": 0,
            "updatedRows": 0,
            "updatedCells": 0,
            "finalRows": len(workbook["rows"]),
            "recordsPath": str(records_path),
            "noOp": True,
        }
        summary_path.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
        return sync_result(
            WRITTEN,
            UNCHANGED,
            "Excel 已包含相同项目数据，未重复新增。",
            matched_by=str(plan.get("matched_by") or ""),
            excel_row=int(record.get("excel_row_id") or 0),
            writer_summary=summary,
        )

    args = Namespace(
        records=str(records_path),
        excel=str(paths.excel),
        sheet_name=paths.sheet_name,
        backup_dir=str(paths.backup_dir),
        summary=str(summary_path),
        override_summary="",
        override_diff="",
        allow_excel_row_id=bool(matched_row),
        append_source_notes=True,
        write_source_url=True,
        overwrite_fields=[],
    )
    summary = run_writer(args)
    if action == INSERT and int(summary.get("appendedRows") or 0) != 1:
        raise RuntimeError("Excel Writer 未按预期新增项目行。")
    if action == UPDATE and int(summary.get("updatedRows") or 0) < 1:
        raise RuntimeError("Excel Writer 未按预期更新已有项目行。")
    return sync_result(
        WRITTEN,
        action,
        {
            INSERT: "项目已新增到 Excel。",
            UPDATE: "Excel 已补充原为空值的字段。",
            UNCHANGED: "Excel 已包含相同项目数据，未重复新增。",
        }[action],
        matched_by=str(plan.get("matched_by") or ""),
        excel_row=int(record.get("excel_row_id") or 0),
        writer_summary=summary,
    )


def sync_confirmed_projects_batch(
    entries: list[dict[str, Any]],
    *,
    paths: InboxExcelSyncPaths,
    batch_id: str = "",
) -> dict[str, dict[str, Any]]:
    """Plan many confirmed projects against one workbook snapshot and save at most once."""
    if not entries:
        return {}
    workbook = inspect_workbook(paths)
    planning_rows = [dict(row) for row in workbook["rows"]]
    audited_by_project, audited_by_detail = prior_excel_row_indexes(paths)
    results: dict[str, dict[str, Any]] = {}
    pending: list[dict[str, Any]] = []
    next_excel_row = len(planning_rows) + 2

    for entry in entries:
        inbox_id = str(entry.get("inbox_id") or "")
        try:
            record = build_excel_record(
                dict(entry.get("extracted") or {}),
                dict(entry.get("project") or {}),
                str(entry.get("source_file") or ""),
                str(entry.get("source_url") or ""),
                str(entry.get("confirmed_time") or ""),
            )
        except Exception as exc:
            results[inbox_id] = sync_result(FAILED, "", str(exc))
            continue
        if not normalize_text(record.get("project_name")):
            results[inbox_id] = sync_result(NOT_APPLICABLE, "", "项目名称缺失，未写入 Excel。")
            continue
        try:
            plan = plan_record_excel_sync(record, planning_rows, audited_by_project, audited_by_detail)
        except Exception as exc:
            results[inbox_id] = sync_result(FAILED, "", str(exc))
            continue
        if plan["excel_status"] == CONFLICT:
            results[inbox_id] = plan
            continue
        matched_row = dict(plan.get("matched_row") or {})
        if matched_row:
            record["excel_row_id"] = int(matched_row["excel_row"])
        action = str(plan["excel_action"])
        predicted_excel_row = int(record.get("excel_row_id") or 0)
        if action == INSERT:
            predicted_excel_row = next_excel_row
            next_excel_row += 1
            planning_rows.append({
                "excel_row": predicted_excel_row,
                "serial": "",
                **{field: record.get(field, "") for field in BUSINESS_FIELDS},
                "note": record.get("note", ""),
                "source_url": record.get("source_url", ""),
                "source_document_id": record.get("source_document_id", ""),
                "award_detail_id": record.get("award_detail_id", ""),
            })
        elif action == UPDATE:
            for row in planning_rows:
                if int(row.get("excel_row") or 0) != predicted_excel_row:
                    continue
                for field in plan.get("updated_fields", []):
                    row[str(field)] = record.get(str(field), "")
                break
        project_id = str(record.get("source_document_id") or "")
        detail_id = str(record.get("award_detail_id") or "")
        if predicted_excel_row >= 2 and action in {INSERT, UPDATE, UNCHANGED}:
            add_index_row(audited_by_project, project_id, predicted_excel_row)
            add_index_row(audited_by_detail, detail_id, predicted_excel_row)

        records_path = paths.records_dir / f"{safe_token(inbox_id)}.json"
        summary_path = paths.summary_dir / f"{safe_token(inbox_id)}.json"
        records_path.parent.mkdir(parents=True, exist_ok=True)
        summary_path.parent.mkdir(parents=True, exist_ok=True)
        records_path.write_text(json.dumps([record], ensure_ascii=False, indent=2), encoding="utf-8")
        if action == UNCHANGED:
            summary = {
                "inputRecords": 1,
                "mergedRecords": 1,
                "existingRows": len(workbook["rows"]),
                "appendedRows": 0,
                "updatedRows": 0,
                "updatedCells": 0,
                "finalRows": len(workbook["rows"]),
                "recordsPath": str(records_path),
                "noOp": True,
            }
            summary_path.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
            results[inbox_id] = sync_result(
                WRITTEN,
                UNCHANGED,
                "Excel 已包含相同项目数据，未重复新增。",
                matched_by=str(plan.get("matched_by") or ""),
                excel_row=predicted_excel_row,
                writer_summary=summary,
            )
            continue
        pending.append({
            "inbox_id": inbox_id,
            "record": record,
            "records_path": records_path,
            "summary_path": summary_path,
            "plan": plan,
            "action": action,
            "predicted_excel_row": predicted_excel_row,
        })

    if not pending:
        return results

    token = safe_token(batch_id or f"batch_{uuid.uuid4().hex}")
    batch_records_path = paths.records_dir / f"{token}.json"
    batch_summary_path = paths.summary_dir / f"{token}.json"
    batch_records_path.write_text(
        json.dumps([dict(item["record"]) for item in pending], ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    args = Namespace(
        records=str(batch_records_path),
        excel=str(paths.excel),
        sheet_name=paths.sheet_name,
        backup_dir=str(paths.backup_dir),
        summary=str(batch_summary_path),
        override_summary="",
        override_diff="",
        allow_excel_row_id=any(bool(item["record"].get("excel_row_id")) for item in pending),
        append_source_notes=True,
        write_source_url=True,
        overwrite_fields=[],
    )
    try:
        summary = run_writer(args)
        detailed_outcomes = summary.get("record_outcomes")
        if not isinstance(detailed_outcomes, list):
            expected_inserts = sum(1 for item in pending if item["action"] == INSERT)
            if int(summary.get("appendedRows") or 0) != expected_inserts:
                raise RuntimeError("Excel Writer 未按批量计划新增中标明细行。")
            expected_updates = sum(1 for item in pending if item["action"] == UPDATE)
            if expected_updates and int(summary.get("updatedRows") or 0) < 1:
                raise RuntimeError("Excel Writer 未按批量计划更新中标明细行。")
    except Exception as exc:
        reason = str(exc)
        for item in pending:
            results[str(item["inbox_id"])] = sync_result(
                FAILED,
                str(item["action"]),
                reason,
                matched_by=str(item["plan"].get("matched_by") or ""),
                excel_row=int(item["predicted_excel_row"] or 0),
            )
        return results

    outcome_rows = detailed_outcomes if isinstance(detailed_outcomes, list) else []
    outcomes_by_detail = {
        str(outcome.get("award_detail_id") or ""): dict(outcome)
        for outcome in outcome_rows
        if isinstance(outcome, dict) and str(outcome.get("award_detail_id") or "")
    }
    outcomes_by_project = {
        str(outcome.get("source_document_id") or ""): dict(outcome)
        for outcome in outcome_rows
        if isinstance(outcome, dict)
        and not str(outcome.get("award_detail_id") or "")
        and str(outcome.get("source_document_id") or "")
    }
    for item in pending:
        inbox_id = str(item["inbox_id"])
        action = str(item["action"])
        detail_id = str(item["record"].get("award_detail_id") or "")
        project_id = str(item["record"].get("source_document_id") or "")
        outcome = outcomes_by_detail.get(detail_id) if detail_id else outcomes_by_project.get(project_id)
        if detail_id and isinstance(detailed_outcomes, list) and (
            outcome is None or str(outcome.get("status") or "") != "written"
        ):
            reason = str((outcome or {}).get("reason") or "Excel Writer 未返回该中标明细的成功结果。")
            results[inbox_id] = sync_result(
                FAILED, action, reason,
                matched_by=str(item["plan"].get("matched_by") or ""),
                excel_row=int((outcome or {}).get("excel_row") or 0),
            )
            continue
        if not detail_id and outcome is not None and str(outcome.get("status") or "") != "written":
            reason = str(outcome.get("reason") or "Excel Writer 未返回该项目的成功结果。")
            results[inbox_id] = sync_result(
                FAILED, action, reason,
                matched_by=str(item["plan"].get("matched_by") or ""),
                excel_row=int(outcome.get("excel_row") or 0),
            )
            continue
        if outcome:
            action = str(outcome.get("action") or action)
        excel_row = int((outcome or {}).get("excel_row") or item["predicted_excel_row"] or 0)
        audit_record = dict(item["record"])
        audit_record["excel_row_id"] = excel_row
        Path(item["records_path"]).write_text(
            json.dumps([audit_record], ensure_ascii=False, indent=2), encoding="utf-8"
        )
        item_summary = {
            "inputRecords": 1,
            "mergedRecords": 1,
            "existingRows": excel_row - 2 if action == INSERT else len(workbook["rows"]),
            "appendedRows": 1 if action == INSERT else 0,
            "updatedRows": 1 if action == UPDATE else 0,
            "finalRows": int(summary.get("finalRows") or len(planning_rows)),
            "excelRow": excel_row,
            "recordsPath": str(item["records_path"]),
            "batchSummaryPath": str(batch_summary_path),
        }
        Path(item["summary_path"]).write_text(
            json.dumps(item_summary, ensure_ascii=False, indent=2), encoding="utf-8"
        )
        results[inbox_id] = sync_result(
            WRITTEN,
            action,
            "中标明细已新增到 Excel。" if action == INSERT else (
                "Excel 已包含该中标明细。" if action == UNCHANGED else "Excel 已补充原为空值的字段。"
            ),
            matched_by=str(item["plan"].get("matched_by") or ""),
            excel_row=excel_row,
            writer_summary=item_summary,
        )
    return results


def build_excel_record(
    extracted: dict[str, Any],
    project: dict[str, Any],
    source_file: str,
    source_url: str,
    confirmed_time: str,
) -> dict[str, Any]:
    trace = dict(project.get("source_trace") or {})
    project_fields = dict(trace.get("extracted_fields") or {})

    def value(key: str, *aliases: str) -> str:
        for field in (key, *aliases):
            candidate = normalize_text(extracted.get(field)) or normalize_text(project_fields.get(field))
            if candidate:
                return candidate
        return ""

    sequence = business_sequence(extracted) or business_sequence(project)
    detail_id = award_detail_id(extracted) or str(project.get("award_detail_id") or "")
    source = normalize_text(source_url) or Path(source_file).name or normalize_text(source_file)
    timestamp = format_local_timestamp(confirmed_time or utc_now())
    return {
        "source_document_id": str(project.get("project_id") or ""),
        "business_sequence": sequence,
        "business_group_id": business_group_id(extracted) or business_group_id(project),
        "award_detail_id": detail_id,
        "source_file": source,
        "source_url": normalize_text(source_url),
        "customer": value("customer"),
        "project_name": value("project_name") or normalize_text(project.get("project_name")),
        "project_number": validated_project_number(value("project_number")),
        "content": value("content"),
        "budget": value("budget"),
        "bid_open_time": value("bid_open_time"),
        "winner": value("winner", "winner_company"),
        "award_amount": value("award_amount"),
        "doc_type": value("doc_type"),
        "note": f"{'业务序号：' + sequence + '；' if sequence else ''}来源：{source or '系统资产库'}；确认时间：{timestamp}",
        "text_chars": 1,
    }


def plan_record_excel_sync(
    record: dict[str, Any],
    rows: list[dict[str, Any]],
    audited_by_project: dict[str, list[int]],
    audited_by_detail: dict[str, list[int]],
) -> dict[str, Any]:
    """Select the legacy project-row rule or the award-detail row rule."""
    project_id = str(record.get("source_document_id") or "")
    detail_id = str(record.get("award_detail_id") or "")
    if not detail_id:
        audited_rows = audited_by_project.get(project_id, [])
        if len(audited_rows) > 1:
            return sync_result(
                CONFLICT, CONFLICT_ACTION,
                "同一资产库项目存在多个 Excel 行审计记录，已停止自动写入。",
                matched_by="repository_project_audit", candidate_rows=audited_rows,
            )
        return plan_excel_sync(record, rows, preferred_excel_row=audited_rows[0] if audited_rows else None)

    detail_rows = audited_by_detail.get(detail_id, [])
    if len(detail_rows) > 1:
        return sync_result(
            CONFLICT, CONFLICT_ACTION,
            "同一中标明细存在多个 Excel 行审计记录，已停止自动写入。",
            matched_by="award_detail_audit", candidate_rows=detail_rows,
        )
    if detail_rows:
        plan = plan_excel_sync(record, rows, preferred_excel_row=detail_rows[0])
        if str(plan.get("matched_by") or "") == "repository_project_audit":
            plan["matched_by"] = "award_detail_audit"
        return plan

    candidate_numbers = set(audited_by_project.get(project_id, []))
    sequence = normalize_text(record.get("business_sequence"))
    if sequence:
        marker = f"业务序号：{sequence}；"
        candidate_numbers.update(
            int(row.get("excel_row") or 0) for row in rows if marker in normalize_text(row.get("note"))
        )
    candidates = [row for row in rows if int(row.get("excel_row") or 0) in candidate_numbers]
    winner = normalize_key(record.get("winner"))
    amount = normalize_key(record.get("award_amount"))
    same_detail = [
        row for row in candidates
        if (winner or amount)
        and normalize_key(row.get("winner")) == winner
        and normalize_key(row.get("award_amount")) == amount
    ]
    if len(same_detail) > 1:
        return sync_result(
            CONFLICT, CONFLICT_ACTION,
            "相同中标单位和金额对应多个 Excel 行，已停止自动写入。",
            matched_by="award_detail_fields_exact",
            candidate_rows=[int(row.get("excel_row") or 0) for row in same_detail],
        )
    if same_detail:
        plan = plan_excel_sync(record, rows, preferred_excel_row=int(same_detail[0]["excel_row"]))
        if str(plan.get("matched_by") or "") == "repository_project_audit":
            plan["matched_by"] = "award_detail_fields_exact"
        return plan

    conflicts: list[dict[str, str]] = []
    for row in candidates:
        for field in PROJECT_LEVEL_FIELDS:
            current = normalize_text(row.get(field))
            incoming = normalize_text(record.get(field))
            if current and incoming and current != incoming:
                conflict = {"field": field, "existing": current, "incoming": incoming}
                if conflict not in conflicts:
                    conflicts.append(conflict)
    if conflicts:
        return sync_result(
            CONFLICT, CONFLICT_ACTION,
            "同项目已有非空项目级字段与确认数据冲突，未新增明细行。",
            matched_by="business_group_project_fields", conflicts=conflicts,
            candidate_rows=[int(row.get("excel_row") or 0) for row in candidates],
        )
    return sync_result(NOT_WRITTEN, INSERT, "将新增 Excel 中标明细行。", matched_by="new_award_detail")


def plan_excel_sync(
    record: dict[str, Any],
    rows: list[dict[str, Any]],
    *,
    preferred_excel_row: int | None = None,
) -> dict[str, Any]:
    number = validated_project_number(record.get("project_number"))
    customer_key = normalize_key(record.get("customer"))
    project_key = normalize_key(record.get("project_name"))
    source_url = normalize_text(record.get("source_url"))

    if preferred_excel_row is not None:
        matches = [row for row in rows if int(row.get("excel_row") or 0) == preferred_excel_row]
        matched_by = "repository_project_audit"
    else:
        source_matches = [
            row for row in rows
            if source_url and (
                normalize_text(row.get("source_url")) == source_url
                or f"来源：{source_url}" in normalize_text(row.get("note"))
            )
        ]
        number_matches: list[dict[str, Any]] = []
        if source_matches:
            matches = source_matches
            matched_by = "source_url_exact"
        elif number:
            target = number.upper()
            for row in rows:
                candidates = project_numbers({
                    "project_name": row.get("project_name", ""),
                    "content": row.get("content", ""),
                    "note": row.get("note", ""),
                    "source_file": "",
                })
                if target in candidates:
                    number_matches.append(row)
            if number_matches:
                matches = number_matches
                matched_by = "project_number_exact"
            else:
                matches = [
                    row for row in rows
                    if customer_key and project_key
                    and normalize_key(row.get("customer")) == customer_key
                    and normalize_key(row.get("project_name")) == project_key
                ]
                matched_by = "customer_project_name_exact"
        else:
            matches = [
                row for row in rows
                if customer_key and project_key
                and normalize_key(row.get("customer")) == customer_key
                and normalize_key(row.get("project_name")) == project_key
            ]
            matched_by = "customer_project_name_exact"

    if len(matches) > 1:
        return sync_result(
            CONFLICT,
            CONFLICT_ACTION,
            "存在多个 Excel 项目候选，已停止自动写入。",
            matched_by=matched_by,
            candidate_rows=[int(row.get("excel_row") or 0) for row in matches],
        )
    if not matches and preferred_excel_row is not None:
        return sync_result(
            CONFLICT,
            CONFLICT_ACTION,
            "资产库项目对应的 Excel 行已不存在，已停止自动写入。",
            matched_by=matched_by,
            candidate_rows=[preferred_excel_row],
        )
    if not matches:
        return sync_result(NOT_WRITTEN, INSERT, "将新增 Excel 项目行。", matched_by="new_project")

    row = dict(matches[0])
    conflicts: list[dict[str, str]] = []
    updates: list[str] = []
    for field in BUSINESS_FIELDS:
        current = normalize_text(row.get(field))
        incoming = normalize_text(record.get(field))
        if current and incoming and current != incoming:
            conflicts.append({"field": field, "existing": current, "incoming": incoming})
        elif not current and incoming:
            updates.append(field)
    if conflicts:
        return sync_result(
            CONFLICT,
            CONFLICT_ACTION,
            "Excel 已有非空值与确认数据冲突，未覆盖原值。",
            matched_by=matched_by,
            excel_row=int(row.get("excel_row") or 0),
            conflicts=conflicts,
        )
    return sync_result(
        NOT_WRITTEN,
        UPDATE if updates else UNCHANGED,
        "将补充 Excel 空字段。" if updates else "Excel 项目数据未变化。",
        matched_by=matched_by,
        excel_row=int(row.get("excel_row") or 0),
        matched_row=row,
        updated_fields=updates,
    )


def prior_project_excel_rows(paths: InboxExcelSyncPaths, project_id: str) -> list[int]:
    """Recover successful project-to-row links from existing per-Inbox writer audits."""
    return prior_project_excel_row_index(paths).get(project_id, []) if project_id else []


def prior_project_excel_row_index(paths: InboxExcelSyncPaths) -> dict[str, list[int]]:
    """Load all successful project-to-row audit links once for batch planning."""
    return prior_excel_row_indexes(paths)[0]


def prior_excel_row_indexes(
    paths: InboxExcelSyncPaths,
) -> tuple[dict[str, list[int]], dict[str, list[int]]]:
    """Load successful project and award-detail row audit links in one scan."""
    if not paths.records_dir.exists() or not paths.summary_dir.exists():
        return {}, {}
    excel_rows: dict[str, set[int]] = {}
    detail_rows: dict[str, set[int]] = {}
    for records_path in paths.records_dir.glob("*.json"):
        try:
            records = json.loads(records_path.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            continue
        if not isinstance(records, list):
            continue
        summary_path = paths.summary_dir / records_path.name
        if not summary_path.exists():
            continue
        try:
            summary = json.loads(summary_path.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            continue
        appended_index = 0
        appended_count = int(summary.get("appendedRows") or 0)
        appended_start = int(summary.get("existingRows") or 0) + 2
        for record in records:
            if not isinstance(record, dict):
                continue
            project_id = str(record.get("source_document_id") or "")
            detail_id = str(record.get("award_detail_id") or "")
            if not project_id and not detail_id:
                continue
            explicit_row = int(record.get("excel_row_id") or 0)
            if explicit_row >= 2:
                if project_id:
                    excel_rows.setdefault(project_id, set()).add(explicit_row)
                if detail_id:
                    detail_rows.setdefault(detail_id, set()).add(explicit_row)
            elif appended_index < appended_count:
                appended_row = appended_start + appended_index
                appended_index += 1
                if appended_row >= 2:
                    if project_id:
                        excel_rows.setdefault(project_id, set()).add(appended_row)
                    if detail_id:
                        detail_rows.setdefault(detail_id, set()).add(appended_row)
    return (
        {project_id: sorted(rows) for project_id, rows in excel_rows.items()},
        {detail_id: sorted(rows) for detail_id, rows in detail_rows.items()},
    )


def add_index_row(index: dict[str, list[int]], identity: str, excel_row: int) -> None:
    if not identity or excel_row < 2:
        return
    rows = index.setdefault(identity, [])
    if excel_row not in rows:
        rows.append(excel_row)
        rows.sort()


def inspect_workbook(paths: InboxExcelSyncPaths) -> dict[str, Any]:
    if not paths.excel.exists():
        raise FileNotFoundError(f"Excel file not found: {paths.excel}")
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("处理 XLSX 需要安装 openpyxl，请先安装项目依赖。") from exc
    workbook = load_workbook(paths.excel, read_only=True, data_only=False)
    try:
        if paths.sheet_name not in workbook.sheetnames:
            raise ValueError(f"Excel sheet not found: {paths.sheet_name}")
        values = [list(row) for row in workbook[paths.sheet_name].iter_rows(values_only=True)]
    finally:
        workbook.close()
    while values and not any(normalize_text(value) for value in values[-1]):
        values.pop()
    headers = [normalize_text(value) for value in (values[0] if values else [])[:10]]
    expected = ["序号", "客户", "项目名称", "项目内容", "预算", "开标时间", "中标厂商", "中标金额", "备注", "来源链接"]
    if headers != expected:
        raise ValueError(f"Unexpected Excel headers: {headers}")
    rows = []
    for excel_row, values_row in enumerate(values[1:], 2):
        padded = values_row + [""] * max(0, 10 - len(values_row))
        rows.append({
            "excel_row": excel_row,
            "serial": padded[0] or "",
            "customer": padded[1] or "",
            "project_name": padded[2] or "",
            "content": padded[3] or "",
            "budget": padded[4] or "",
            "bid_open_time": padded[5] or "",
            "winner": padded[6] or "",
            "award_amount": padded[7] or "",
            "note": padded[8] or "",
            "source_url": padded[9] or "",
        })
    return {"headers": headers, "rows": rows}


def repair_confirmed_project_excel(
    *,
    audit_id: str,
    extracted: dict[str, Any],
    project: dict[str, Any],
    source_url: str,
    confirmed_time: str,
    expected_serial: int,
    expected_current_project_name: str,
    paths: InboxExcelSyncPaths,
) -> dict[str, Any]:
    """Apply an explicitly authorized correction after unique URL and identity checks."""
    record = build_excel_record(extracted, project, "", source_url, confirmed_time)
    workbook = inspect_workbook(paths)
    url_matches = [
        row for row in workbook["rows"]
        if normalize_text(row.get("source_url")) == source_url
        or f"来源：{source_url}" in normalize_text(row.get("note"))
    ]
    if len(url_matches) != 1:
        return sync_result(
            CONFLICT, CONFLICT_ACTION,
            "来源 URL 未唯一定位到 Excel 记录，已停止修正。",
            matched_by="source_url_exact", candidate_rows=[int(row.get("excel_row") or 0) for row in url_matches],
        )
    row = dict(url_matches[0])
    number = validated_project_number(record.get("project_number"))
    candidates = project_numbers({
        "project_name": row.get("project_name", ""), "content": row.get("content", ""),
        "note": row.get("note", ""), "source_file": "",
    })
    if (
        int(row.get("serial") or 0) != expected_serial
        or normalize_text(row.get("project_name")) != normalize_text(expected_current_project_name)
        or not number
        or number.upper() not in candidates
    ):
        return sync_result(
            CONFLICT, CONFLICT_ACTION,
            "URL 命中行未通过项目编号、项目序号和当前项目名称交叉核验，已停止修正。",
            matched_by="source_url_exact", candidate_rows=[int(row.get("excel_row") or 0)],
        )

    record["excel_row_id"] = int(row["excel_row"])
    records_path = paths.records_dir / f"{safe_token(audit_id)}.json"
    summary_path = paths.summary_dir / f"{safe_token(audit_id)}.json"
    records_path.parent.mkdir(parents=True, exist_ok=True)
    summary_path.parent.mkdir(parents=True, exist_ok=True)
    records_path.write_text(json.dumps([record], ensure_ascii=False, indent=2), encoding="utf-8")
    args = Namespace(
        records=str(records_path), excel=str(paths.excel), sheet_name=paths.sheet_name,
        backup_dir=str(paths.backup_dir), summary=str(summary_path),
        override_summary="", override_diff="",
        allow_excel_row_id=True, append_source_notes=True, write_source_url=True,
        overwrite_fields=["customer", "project_name", "content", "winner", "award_amount"],
    )
    summary = run_writer(args)
    if int(summary.get("updatedRows") or 0) != 1:
        raise RuntimeError("Excel Writer 未按预期修正唯一目标行。")
    return sync_result(
        WRITTEN, UPDATE, "Excel 目标项目行已完成授权修正。",
        matched_by="source_url_exact", excel_row=int(row["excel_row"]), writer_summary=summary,
    )


def sync_result(status: str, action: str, message: str, **extra: Any) -> dict[str, Any]:
    return {"excel_status": status, "excel_action": action, "message": message, **extra}


def normalize_text(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def normalize_key(value: object) -> str:
    normalized = unicodedata.normalize("NFKC", normalize_text(value)).lower()
    return re.sub(r"[\s（）()【】\[\]《》<>“”\"':：;；,，.。_\-—+]", "", normalized)


def safe_token(value: str) -> str:
    return re.sub(r"[^A-Za-z0-9_.-]+", "_", value)[:120] or uuid.uuid4().hex


def format_local_timestamp(value: str, target_timezone: Any | None = None) -> str:
    """Render an audit timestamp in the service OS timezone without changing stored UTC audit values."""
    raw = str(value or "").strip()
    if not raw:
        raw = utc_now()
    parsed = datetime.fromisoformat(raw[:-1] + "+00:00" if raw.endswith("Z") else raw)
    if parsed.tzinfo is None:
        parsed = parsed.replace(tzinfo=timezone.utc)
    localized = parsed.astimezone(target_timezone)
    offset = localized.strftime("%z")
    formatted_offset = f"{offset[:3]}:{offset[3:]}" if len(offset) == 5 else "+00:00"
    return f"{localized.strftime('%Y-%m-%d %H:%M:%S')} UTC{formatted_offset}"


def utc_now() -> str:
    return datetime.now(timezone.utc).isoformat(timespec="seconds").replace("+00:00", "Z")
