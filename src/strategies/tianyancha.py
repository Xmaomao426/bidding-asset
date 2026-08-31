from __future__ import annotations

import hashlib
import json
import re
from collections import Counter
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


HEADER_FIELDS = [
    "序号",
    "发布日期",
    "标题",
    "省份地区",
    "公告类型",
    "招采单位",
    "中标单位",
    "中标金额",
    "招投标详情",
]

EMPTY_MARKS = {"", "-", "未明确列出", "None", "null"}
AWARD_NOTICE_MARKS = ("中标", "成交", "合同")
NORMAL_NOTICE_MARKS = ("招标", "中标", "成交", "合同")
WEAK_KEYWORDS = {"AI", "信息化", "数字化", "智能化"}


def load_json(path: Path, default: Any) -> Any:
    if not path.exists():
        return default
    return json.loads(path.read_text(encoding="utf-8"))


def write_json(payload: Any, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def clean(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def is_empty(value: Any) -> bool:
    return clean(value) in EMPTY_MARKS


def normalize_text(value: Any) -> str:
    text = clean(value).lower()
    text = re.sub(r"\s+", "", text)
    return re.sub(r"[^\w\u4e00-\u9fff]+", "", text)


def stable_id(*parts: Any) -> str:
    raw = "|".join(clean(part) for part in parts)
    return "tyc_" + hashlib.sha256(raw.encode("utf-8")).hexdigest()[:16]


def is_award_notice(notice_type: str) -> bool:
    text = clean(notice_type)
    return any(mark in text for mark in AWARD_NOTICE_MARKS)


def is_normal_notice(notice_type: str) -> bool:
    text = clean(notice_type)
    return any(mark in text for mark in NORMAL_NOTICE_MARKS)


def expand_export_paths(exports: list[str]) -> list[Path]:
    paths: list[Path] = []
    for item in exports:
        pattern = Path(item)
        if any(ch in item for ch in "*?[]"):
            parent = pattern.parent if str(pattern.parent) else Path(".")
            matches = sorted(parent.glob(pattern.name))
            paths.extend(
                match for match in matches if match.suffix.lower() == ".xlsx" and not match.name.startswith("~$")
            )
        else:
            if not pattern.name.startswith("~$"):
                paths.append(pattern)
    seen: set[str] = set()
    unique: list[Path] = []
    for path in paths:
        resolved = str(path.resolve())
        if resolved not in seen:
            seen.add(resolved)
            unique.append(path)
    return unique


def find_header_row(ws: Any) -> tuple[int, dict[str, int]]:
    for row_number in range(1, min(ws.max_row, 10) + 1):
        values = [clean(ws.cell(row_number, col).value) for col in range(1, ws.max_column + 1)]
        if all(field in values for field in HEADER_FIELDS):
            return row_number, {
                clean(ws.cell(row_number, col).value): col for col in range(1, ws.max_column + 1)
            }
    raise ValueError(f"Could not find Tianyancha header in sheet {ws.title!r}")


def merge_continuation(current: dict[str, Any], winner: str, amount: str) -> None:
    if not is_empty(winner):
        winners = [item for item in current["winner"].split("；") if item]
        if winner not in winners:
            winners.append(winner)
        current["winner"] = "；".join(winners)
    if not is_empty(amount):
        amounts = [item for item in current["award_amount"].split("；") if item]
        if amount not in amounts:
            amounts.append(amount)
        current["award_amount"] = "；".join(amounts)


def parse_export(path: Path) -> dict[str, Any]:
    workbook = load_workbook(path, read_only=False, data_only=True)
    worksheet = workbook.active
    header_row, headers = find_header_row(worksheet)
    rows: list[dict[str, Any]] = []
    current: dict[str, Any] | None = None
    continuation_rows = 0
    rows_with_hyperlink = 0

    for row_number in range(header_row + 1, worksheet.max_row + 1):
        sequence = clean(worksheet.cell(row_number, headers["序号"]).value)
        title = clean(worksheet.cell(row_number, headers["标题"]).value)
        winner = clean(worksheet.cell(row_number, headers["中标单位"]).value)
        amount = clean(worksheet.cell(row_number, headers["中标金额"]).value)
        detail_cell = worksheet.cell(row_number, headers["招投标详情"])
        source_url = detail_cell.hyperlink.target if detail_cell.hyperlink else ""

        if not sequence and not title and current and (winner or amount):
            continuation_rows += 1
            merge_continuation(current, winner, amount)
            continue
        if not sequence and not title and not source_url:
            continue
        if source_url:
            rows_with_hyperlink += 1

        row = {
            "row_id": stable_id(path.name, sequence, title, source_url),
            "source": "tianyancha_export",
            "source_file": path.name,
            "row_number": row_number,
            "sequence": sequence,
            "publish_date": clean(worksheet.cell(row_number, headers["发布日期"]).value),
            "title": title,
            "project_name": title,
            "region": clean(worksheet.cell(row_number, headers["省份地区"]).value),
            "notice_type": clean(worksheet.cell(row_number, headers["公告类型"]).value),
            "customer": clean(worksheet.cell(row_number, headers["招采单位"]).value),
            "winner": "" if is_empty(winner) else winner,
            "award_amount": "" if is_empty(amount) else amount,
            "source_url": source_url,
        }
        rows.append(row)
        current = row

    return {
        "file": str(path),
        "file_name": path.name,
        "sheet": worksheet.title,
        "max_row": worksheet.max_row,
        "max_col": worksheet.max_column,
        "header_row": header_row,
        "parsed_rows": len(rows),
        "rows_with_hyperlink": rows_with_hyperlink,
        "continuation_rows_merged": continuation_rows,
        "notice_type_counts": dict(Counter(row["notice_type"] for row in rows)),
        "rows": rows,
    }


def load_keywords(path: Path) -> dict[str, Any]:
    config = load_json(path, {})
    return {
        "enabled": bool(config.get("enabled", True)),
        "match_mode": config.get("match_mode", "any"),
        "title_keywords": [clean(item) for item in config.get("title_keywords", []) if clean(item)],
        "exclude_keywords": [clean(item) for item in config.get("exclude_keywords", []) if clean(item)],
    }


def keyword_matches(title: str, keywords: list[str], mode: str) -> list[str]:
    title_norm = normalize_text(title)
    matches = []
    for keyword in keywords:
        keyword_norm = normalize_text(keyword)
        if keyword_norm and keyword_norm in title_norm:
            matches.append(keyword)
    if mode == "all" and len(matches) != len(keywords):
        return []
    return matches


def build_context_index(final_records_path: Path, review_queue_path: Path) -> list[dict[str, str]]:
    index: list[dict[str, str]] = []
    for record in load_json(final_records_path, []):
        if not isinstance(record, dict):
            continue
        for field in ("project_name", "content"):
            value = clean(record.get(field))
            normalized = normalize_text(value)
            if normalized:
                index.append(
                    {
                        "source": "final_records",
                        "record_id": clean(record.get("source_document_id")) or clean(record.get("record_id")),
                        "source_file": clean(record.get("source_file")),
                        "project_name": clean(record.get("project_name")),
                        "normalized": normalized,
                    }
                )
    for item in load_json(review_queue_path, []):
        if not isinstance(item, dict):
            continue
        value = clean(item.get("project_name"))
        normalized = normalize_text(value)
        if normalized:
            index.append(
                {
                    "source": "manual_review_queue",
                    "record_id": clean(item.get("record_id")) or clean(item.get("source_document_id")),
                    "source_file": clean(item.get("source_file")),
                    "project_name": value,
                    "normalized": normalized,
                }
            )
    return index


def match_existing(row: dict[str, Any], context_index: list[dict[str, str]]) -> list[dict[str, str]]:
    title_norm = normalize_text(row.get("title"))
    if not title_norm:
        return []
    matches = []
    for item in context_index:
        existing = item["normalized"]
        if title_norm == existing or (len(title_norm) >= 12 and title_norm in existing) or (
            len(existing) >= 12 and existing in title_norm
        ):
            matches.append({key: item[key] for key in ("source", "record_id", "source_file", "project_name")})
    return matches


def risk_flags_for(row: dict[str, Any], matched_keywords: list[str], is_new_record: bool) -> list[str]:
    flags = ["third_party_source"]
    if is_new_record:
        flags.append("new_record_candidate")
    if not row.get("source_url"):
        flags.append("source_url_missing")
    if not row.get("publish_date"):
        flags.append("publish_date_missing")
    if not is_normal_notice(row.get("notice_type", "")):
        flags.append("unusual_notice_type")
    if "；" in row.get("winner", "") or "；" in row.get("award_amount", ""):
        flags.extend(["package_unclear", "multiple_winners"])
    if matched_keywords and all(keyword in WEAK_KEYWORDS for keyword in matched_keywords):
        flags.append("weak_keyword_match")
    return flags


def suggested_excel_fields(row: dict[str, Any]) -> dict[str, str]:
    fields = {
        "project_name": clean(row.get("title")),
        "customer": clean(row.get("customer")),
        "content": clean(row.get("title")),
        "note": "来源：天眼查导出，关键词命中，待人工确认",
    }
    if is_award_notice(row.get("notice_type", "")):
        if not is_empty(row.get("winner")):
            fields["winner"] = clean(row.get("winner"))
        if not is_empty(row.get("award_amount")):
            fields["award_amount"] = clean(row.get("award_amount"))
    return {key: value for key, value in fields.items() if value}


def build_candidate(
    row: dict[str, Any],
    matched_keywords: list[str],
    exclude_matches: list[str],
    existing_matches: list[dict[str, str]],
    is_new_record: bool,
) -> dict[str, Any]:
    return {
        "candidate_id": stable_id(row.get("row_id"), "new" if is_new_record else "matched"),
        "row_id": row.get("row_id", ""),
        "source": "tianyancha_export",
        "source_file": row.get("source_file", ""),
        "source_row": row.get("row_number", ""),
        "title": row.get("title", ""),
        "project_name": row.get("project_name", ""),
        "publish_date": row.get("publish_date", ""),
        "region": row.get("region", ""),
        "notice_type": row.get("notice_type", ""),
        "customer": row.get("customer", ""),
        "winner": row.get("winner", ""),
        "award_amount": row.get("award_amount", ""),
        "source_url": row.get("source_url", ""),
        "matched_keywords": matched_keywords,
        "exclude_keywords": exclude_matches,
        "matched_existing_records": existing_matches,
        "risk_flags": risk_flags_for(row, matched_keywords, is_new_record),
        "suggested_excel_fields": suggested_excel_fields(row) if is_new_record else {},
        "review_status": "needs_manual_review",
    }


def write_candidates_markdown(payload: dict[str, Any], path: Path) -> None:
    lines = [
        "# Tianyancha Matched Candidates",
        "",
        "| 序号 | 发布日期 | 公告类型 | 标题 | 招采单位 | 中标单位 | 中标金额 | 匹配记录 | 来源 |",
        "|---:|---|---|---|---|---|---|---|---|",
    ]
    for index, item in enumerate(payload["matched_candidates"], 1):
        matched = "; ".join(match["source"] + ":" + match["record_id"] for match in item["matched_existing_records"])
        source = f"[查看详情]({item['source_url']})" if item.get("source_url") else ""
        lines.append(
            "| {index} | {publish_date} | {notice_type} | {title} | {customer} | {winner} | {amount} | {matched} | {source} |".format(
                index=index,
                publish_date=md(item.get("publish_date")),
                notice_type=md(item.get("notice_type")),
                title=md(item.get("title")),
                customer=md(item.get("customer")),
                winner=md(item.get("winner")),
                amount=md(item.get("award_amount")),
                matched=md(matched),
                source=source,
            )
        )
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def write_new_records_markdown(payload: dict[str, Any], path: Path) -> None:
    lines = [
        "# Tianyancha New Record Candidates",
        "",
        "| 序号 | 发布日期 | 公告类型 | 命中关键词 | 风险标记 | 标题 | 招采单位 | 建议字段 | 来源 |",
        "|---:|---|---|---|---|---|---|---|---|",
    ]
    for index, item in enumerate(payload["new_record_candidates"], 1):
        fields = "; ".join(f"{key}={value}" for key, value in item["suggested_excel_fields"].items())
        source = f"[查看详情]({item['source_url']})" if item.get("source_url") else ""
        lines.append(
            "| {index} | {publish_date} | {notice_type} | {keywords} | {flags} | {title} | {customer} | {fields} | {source} |".format(
                index=index,
                publish_date=md(item.get("publish_date")),
                notice_type=md(item.get("notice_type")),
                keywords=md("；".join(item.get("matched_keywords", []))),
                flags=md("；".join(item.get("risk_flags", []))),
                title=md(item.get("title")),
                customer=md(item.get("customer")),
                fields=md(fields),
                source=source,
            )
        )
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def md(value: Any) -> str:
    return clean(value).replace("|", "\\|").replace("\n", " ")[:240]


def generate(
    exports: list[str],
    keyword_config_path: Path,
    final_records_path: Path,
    review_queue_path: Path,
) -> dict[str, Any]:
    keyword_config = load_keywords(keyword_config_path)
    context_index = build_context_index(final_records_path, review_queue_path)
    parsed_exports = [parse_export(path) for path in expand_export_paths(exports)]

    raw_rows: list[dict[str, Any]] = []
    matched_candidates: list[dict[str, Any]] = []
    new_record_candidates: list[dict[str, Any]] = []
    notice_type_counts: Counter[str] = Counter()
    matched_keyword_counts: Counter[str] = Counter()
    exclude_keyword_counts: Counter[str] = Counter()
    risk_flag_counts: Counter[str] = Counter()
    keyword_matched_rows = 0
    excluded_by_keyword_rows = 0

    for parsed in parsed_exports:
        notice_type_counts.update(parsed["notice_type_counts"])
        for row in parsed["rows"]:
            raw_rows.append(row)
            existing_matches = match_existing(row, context_index)
            title_matches = keyword_matches(
                row.get("title", ""),
                keyword_config["title_keywords"] if keyword_config["enabled"] else [],
                keyword_config["match_mode"],
            )
            exclude_matches = keyword_matches(row.get("title", ""), keyword_config["exclude_keywords"], "any")
            if title_matches:
                keyword_matched_rows += 1
                matched_keyword_counts.update(title_matches)
            if exclude_matches:
                excluded_by_keyword_rows += 1
                exclude_keyword_counts.update(exclude_matches)

            if existing_matches:
                candidate = build_candidate(row, title_matches, exclude_matches, existing_matches, False)
                matched_candidates.append(candidate)
                risk_flag_counts.update(candidate["risk_flags"])
                continue

            if title_matches and not exclude_matches:
                candidate = build_candidate(row, title_matches, exclude_matches, [], True)
                new_record_candidates.append(candidate)
                risk_flag_counts.update(candidate["risk_flags"])

    imported_files = [
        {key: value for key, value in parsed.items() if key != "rows"} for parsed in parsed_exports
    ]
    summary = {
        "generated_at": "2026-07-07",
        "keyword_config_path": str(keyword_config_path),
        "title_keyword_count": len(keyword_config["title_keywords"]),
        "exclude_keyword_count": len(keyword_config["exclude_keywords"]),
        "imported_files": imported_files,
        "total_rows": len(raw_rows),
        "parsed_rows": len(raw_rows),
        "worksheet_rows": sum(item["max_row"] for item in imported_files),
        "rows_with_hyperlink": sum(item["rows_with_hyperlink"] for item in imported_files),
        "matched_candidate_count": len(matched_candidates),
        "candidate_count": len(matched_candidates) + len(new_record_candidates),
        "keyword_matched_rows": keyword_matched_rows,
        "excluded_by_keyword_rows": excluded_by_keyword_rows,
        "new_record_candidate_count": len(new_record_candidates),
        "matched_keyword_counts": dict(matched_keyword_counts),
        "exclude_keyword_counts": dict(exclude_keyword_counts),
        "risk_flag_counts": dict(risk_flag_counts),
        "notice_type_counts": dict(notice_type_counts),
    }
    return {
        "generated_at": summary["generated_at"],
        "source_type": "tianyancha_export",
        "keyword_config": keyword_config,
        "summary": summary,
        "raw_rows": raw_rows,
        "matched_candidates": matched_candidates,
        "new_record_candidates": new_record_candidates,
    }
